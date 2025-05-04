# Required Libraries:
import os
import json
import logging
from typing import Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from urllib.parse import urlparse
from book import Book
from audiblescrapernct.configuration import Configuration #
from audiblescrapernct.load_config import load_config #

logger = logging.getLogger(__name__)

# def ensure_directory_exists(path: str) -> None:
def setup_environment(config: Configuration) -> Tuple[str, str, str]:
    """
    Creates necessary directories (if needed) and returns key file paths.
    
    Args: 
        config: Configuration object containing paths and settings.
    """
    # Configuration.from_dict already makes paths absolute
    logger.info("Setting up environment...")
    json_input_path = os.path.normpath(os.path.join(config.data_folder, config.output_json_file))
    logger.info(f"JSON input path (for reading the previously created json file containing all the items in your library): {json_input_path}")
    excel_output_path = os.path.normpath(os.path.join(config.data_folder, config.output_markdown_file))
    logger.info(f"The Excel file that will be created: {excel_output_path}")
    images_path = os.path.normpath(config.images_folder)
    logger.info(f"Images save path from, where to read images for creating the markdown file: {images_path}")

    try:
        # data_folder is already created by Configuration logic if relative
        os.makedirs(images_path, exist_ok=True)
        logger.info(f"✅ Data directory: '{config.data_folder}'") #
        logger.info(f"✅ Images directory: '{images_path}'") #
    except OSError as e:
        logger.error(f"❌ Error ensuring directories exist: {repr(e)}. Check permissions.")
        raise # Re-raise the error to stop execution
    return json_input_path, images_path, excel_output_path


def read_json_file() -> list[Book]:
    """Reads book data from the predefined JSON file and deserializes it into Book objects.

    Checks for the existence of the JSON file (defined by the global constant
    DESTINATION_JSON_FILE) and the associated images folder (IMAGES_FOLDER)
    before attempting to load and parse the JSON data. Assumes the JSON file
    contains a list of dictionaries that match the structure of the Book dataclass.

    Returns:
        list[Book]: A list of Book objects loaded from the JSON file. Returns
                    an empty list if the JSON file or images folder is not
                    found, or if any error occurs during file reading or JSON
                    parsing.

    Raises:
        - FileNotFoundError: If the DESTINATION_JSON_FILE does not exist.
        - json.JSONDecodeError: If the JSON file is improperly formatted.
        - Other Exceptions: Catches general exceptions during file operations.
        In all handled exception cases, an error message is printed, and an
        empty list is returned.

    Side Effects:
        Prints error messages to standard output if the required files are
        missing or if errors occur during processing.
    """
    try:
        # Check if the JSON file and images folder exist
        # Note: Using global constant instead of file_path argument
        if not os.path.exists(DESTINATION_JSON_FILE):
            print(f"❌ JSON file '{DESTINATION_JSON_FILE}' not found.")
            return [] # Return empty list

        if not os.path.exists(IMAGES_FOLDER):
            print(f"❌ Images folder '{IMAGES_FOLDER}' not found.")
            return [] # Return empty list

        # Load the JSON data
        with open(DESTINATION_JSON_FILE, "r", encoding="utf-8") as f:
            # Assuming the JSON file contains a list of book objects
            # We need to deserialize them into Book objects
            books_data = json.load(f)
            books = [Book(**book_data) for book_data in books_data] # Deserialize into Book objects
        return books
    except FileNotFoundError: # Be more specific with exceptions if possible
        print(f"❌ JSON file '{DESTINATION_JSON_FILE}' not found.")
        return [] # Return empty list on FileNotFoundError
    except json.JSONDecodeError as e:
        print(f"❌ Error decoding JSON from '{DESTINATION_JSON_FILE}': {repr(e)}")
        return [] # Return empty list on JSON error
    except Exception as e:
        print(f"❌ An unexpected error occurred while reading JSON file: {repr(e)}")
        return [] # Return empty list for other errors

def create_excel_workbook(books: list[Book]) -> None:
    """Creates an Excel workbook from a list of Book objects.

    Generates an Excel file containing the details of the provided books,
    including title, author, narrator, series, description, cover URL,
    image filename, and the cover image itself embedded in a cell.
    Applies text wrapping to the description column.

    Args:
        books (list[Book]): A list of Book objects to include in the Excel file.

    Returns:
        None

    Raises:
        - Catches general exceptions during workbook creation, image handling,
            or file saving. Prints an error message if an exception occurs.

    Side Effects:
        - Creates or overwrites an Excel file named according to the
            OUTPUT_EXCEL_FILE constant.
        - Prints a success or error message to standard output.
    """
    if not books: # Add a check if the list is empty
        print("⚠️ No book data provided to create Excel workbook.")
        return
    try:
        # Create a new Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Audible Library"

        # Define the headers for the Excel file
        headers = ["Title", "Author", "Narrator", "Series", "Description", "Cover URL", "Filename", "Cover Image"]
        ws.append(headers)

        # Populate the worksheet with book data
        for idx, book in enumerate(books, start=2):
            # Access attributes directly from the Book object
            title = book.title
            author = book.author
            narrator = book.narrator
            series = book.series
            description = book.description
            cover_url = book.cover_image_url

            filename = os.path.basename(urlparse(cover_url).path) if cover_url else ""

            ws.append([
                title,
                author,
                narrator,
                series,
                description,
                cover_url,
                filename
            ])

            img_path = os.path.join(IMAGES_FOLDER, filename)
            if filename and os.path.exists(img_path): # Check filename exists too
                try:
                    img = OpenpyxlImage(img_path)
                    img.width = 75
                    img.height = 75
                    # Set row height (in points) - adjust value as needed
                    ws.row_dimensions[idx].height = 60
                    # Set column width (in character units) - adjust value as needed
                    ws.column_dimensions['H'].width = 12
                    img_anchor = f"H{idx}"
                    ws.add_image(img, img_anchor)
                except Exception as img_e:
                    print(f"⚠️ Could not add image {filename} for '{title}': {repr(img_e)}")


        # Set text wrapping for the description column and center vertically
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column): # Apply to header too
            for cell in row:
                cell.alignment = Alignment(vertical='top') # Align all cells top initially
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5): # Wrap description
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8): # Center image column
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


        wb.save(OUTPUT_EXCEL_FILE)
        print(f"✅ Excel file '{OUTPUT_EXCEL_FILE}' created successfully with {len(books)} books.")
    except Exception as e:
            print(f"❌ Error generating Excel document: {repr(e)}")

def create_audible_library_excel():
    """Orchestrates the creation of the Audible Library Excel file.

    Calls `read_json_file` to load book data from the predefined JSON file.
    If book data is successfully loaded (i.e., the returned list is not empty),
    it then calls `create_excel_workbook` to generate the Excel file.

    Returns:
        None

    Side Effects:
        - Calls `read_json_file` and `create_excel_workbook`, inheriting their
          side effects (printing status/error messages, file I/O).
        - Prints an error message if `read_json_file` returns no books.
    """
    books: list[Book] = read_json_file() # Removed argument as per previous change
    if not books:
        # read_json_file already prints specific errors, so just a general message here
        print("❌ Exiting: Could not load book data to create Excel file.")
        return
    # Create the Excel workbook with the book data
    create_excel_workbook(books)


# Entry point for the script
if __name__ == "__main__":
    create_audible_library_excel()